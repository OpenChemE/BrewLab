# -*- coding: utf-8 -*-

# Import libraries
from time import sleep
import time
import serial
from openpyxl import Workbook
#from openpyxl.compat import range
import os
import datetime
import matplotlib.pyplot as plt
import keyboard # Using module keyboard
import sys

MODE = "dev"

if MODE is "dev":
    from brewlab import fakeSerial as serial

# FUNCTIONS
# This function asks user if they will be using a fermenter for the experiment
def fermChoose(fermNum,ser):
    ferm_stat = False   
    print("Will fermenter " + str(fermNum) + " be run for the experiment? type y/n for yes or no")  # Ask user which fermenters will be run
    while (True):
        reply = str(input())
        if (reply == "y" or reply == "Y"):
            ferm_stat = True
            print("Activating fermenter " + str(fermNum))
            ser.write('1') #send message to Arduino to log data from fermenter        
            sleep(0.1)
            return ferm_stat
        elif (reply == "n" or reply == "N"):
            # Send message to Arduino to NOT log data from fermenter
            ser.write('0') # Send message to Arduino to log data from fermenter        
            sleep(0.1)
            return ferm_stat
        else:
            "error type y for yes or n for no\n"
    
# This function asks user what temperature they want to set fermenter to
def fermTemp(fermNum):
    print("\nplease input an integer for temperature from 4 - 23 C")
    while (True):    
        reply = int(input())
        if (reply >= 4 and reply <= 23):
            print("\nSetting fermenter temperature to "+ str(reply)+" C\n")
            return float(reply)
        else:
            print("error type an integer between 4 and 23\n")

# This function generates a filename based on the date & test number
# It checks existing files in the folder to ensure that no overwrite occurs            
def createDataFile():
    path = os.chdir("C:\\Users\Public\Documents\CHBreweryData")  # Define directory
    today = datetime.date.today()  # Get current date

    testNum = 1  # Set current test number
    dest_filename = str(today)+ "test" + str(testNum) +".xlsx"  # Generate file name

    fileList = os.listdir(path)  # Get list of items in folder

    lengthList = len(fileList)  # Get list length

    for x in range(0,lengthList):#for the length of the list
        fileName = fileList[x] #get the name of the list
        #print (fileName[0:10])
        if str(fileName[0:10]) == str(today):#if date of file is today
            testNum = testNum + 1#increase test number by 1
            #print ("confirmed")
    dest_filename = str(today)+ "test" + str(testNum) +".xlsx"#create name  
    wb.save(filename = dest_filename)#save workbook
    return dest_filename

# Connects to Arduino and prints hello to confirm connection
def ardCon(COM_NUM):
    print ("Connecting to Arduino\n")
    try:
        ser = serial.Serial(COM_NUM, 9600, timeout=0) # Establish the connection on specified COM port
        sleep(2)
        ser.write("Hello")
        sleep(5)
        reply = ser.readline()
        print(reply.strip('\n'))
    except AttributeError as e:
        #serial.close()
        print(e)
        print ("No connection to Arduino, terminating program")
        sys.exit()
    return ser

# Begin communication with arduinos connected to the system
COM1 = 'COM3'
COM2 = 'COM4'
COM3 = 'COM5'

ser1 = ardCon(COM1)
ser2 = ardCon(COM2)
ser3 = ardCon(COM3)

# Ask users which fermenters they want to use
ferm1 = fermChoose(1,ser1)
if (ferm1 == True):
    fTemp1 = fermTemp(1)
ferm2 = fermChoose(2,ser2)
if (ferm2 == True):
    fTemp2 = fermTemp(2)
ferm3 = fermChoose(3,ser3)
if (ferm3 == True):
    fTemp3 = fermTemp(3)

# Create workbook
wb = Workbook()  # generate workbook
ws1 = wb.active  # activate workbook
ws1.title = "Fermentation Data"

fileName = createDataFile()
print (str(fileName) + " data file generated")
    
# Create lists to hold data
if (ferm1 == True):
    T1 = list()
    PH1 = list()
    DO1 = list()
    t1 = list()
    ws1['A1'] = "Fermenter 1 Data"
    # define data entry headers
    ws1['A2'] = "Time (min)"
    ws1['B2'] = "Temperature (C)"
    ws1['C2'] = "PH"
    ws1['D2'] = "Dissolved Oxygen (mg/L)"
    
if (ferm2 == True):
    T2 = list()
    PH2 = list()
    DO2 = list()
    t2 = list()
    ws1['E1'] = "Fermenter 2 Data"
    # define data entry headers
    ws1['E2'] = "Time (min)"
    ws1['F2'] = "Temperature (C)"
    ws1['G2'] = "PH"
    ws1['H2'] = "Dissolved Oxygen (mg/L)"
    
if (ferm3 == True):
    T3 = list()
    PH3 = list()
    DO3 = list()
    t3 =list()
    ws1['I1'] = "Fermenter 3 Data"
    # define data entry headers
    ws1['I2'] = "Time (min)"
    ws1['J2'] = "Temperature (C)"
    ws1['K2'] = "PH"
    ws1['L2'] = "Dissolved Oxygen (mg/L)"
    
# Use infinite loop for data acquisition
fermAct = True
sleep(0.5)  # Wait a bit for arduino to catch up

# Define flags
PHflag = False
DOflag = False
tflag = False
Tflag = False
figure1Active = False
figure2Active = False
figure3Active = False
count1 = 0
count2 = 0
count3 = 0
i1 = 0#counter for excel sheet
i2 = 0
i3 = 0
saveCount = 0#counter for saving excel sheet
dataRun = False

while (fermAct == True):

    if (ferm1 == True):
        ser1.write('F\n')  # let arduino know to send F1 data
        time.sleep(0.5)
        data = ser1.readline()  # readline from buffer and remove newline
        print(data)
        if (str(data[0:2]) == "F1"):
            time.sleep(0.1)
            count = 0  # reate counter for data collection
            while (count < 4):  # while all data has not been collected
                data = ser1.readline()  # readline from buffer and remove newline
                print(data)
            
                if (str(data[0:2]) =="ti" and tflag == False):
                    t1.append(float(data[2:]))
                    count = count + 1
                    tflag = True
            
                if (str(data[0:2]) == "PH" and PHflag == False):
                    try:
                        strCount = 0
                        for x in range (0,len(data)):
                            check = data[x]
                            if (check == "."):
                                strCount = strCount + 1
                                strPlace = x
                        if (strCount > 1):
                            PH1.append(float((data[(x-1):]).strip('\r')))
                        else:
                            PH1.append(float((data[2:]).strip('\r')))
                    except:
                        PH1.append(len(PH1)-1)
                        #PH1.append(float((data[2:5]).strip('\r')))
                    count = count + 1
                    PHflag = True
                
                if (str(data[0:2]) == "DO" and DOflag == False):
                    try:
                        DO1.append(float(data[2:]))
                    except:
                        DO1.append(0)
                    count = count + 1
                    DOflag = True
                
                if (str(data[0:2]) == "Te" and Tflag == False):
                    T1.append(float(data[2:]))
                    count = count + 1
                    Tflag = True
                
            #Reset Flags
            PHflag = False
            DOflag = False
            tflag = False
            Tflag = False
            
            #Record Data into Excel file
            d = ws1.cell(row = 2+i1, column = 1)#call cell
            d.value = t1[i1] #set cell value
            d = ws1.cell(row = 2+i1, column = 2)#call cell
            d.value = T1[i1] #set cell value
            d = ws1.cell(row = 2+i1, column = 3)#call cell
            d.value = PH1[i1] #set cell value
            d = ws1.cell(row = 2+i1, column = 4)#call cell
            d.value = DO1[i1] #set cell value
            i1 = i1 + 1
            saveCount = saveCount + 1
            if (saveCount == 10):
                wb.save(filename = fileName)
                saveCount = 0
            
            if (figure1Active == False):#define first figure
                plt.ion()#create interactive
                plt.figure(1)
                plt.subplot(311)#create plot of time vs. Temp    
                plt.title('Fermenter 1')
                plt.plot(t1,T1, color = 'r')
                plt.ylim(0,30)
                plt.xlabel('Time (min)')
                plt.ylabel('Temp (*C)')
            
                plt.subplot(312)#create plot of time vs. PH
                plt.plot(t1,PH1, color = 'b')
                plt.ylim(0,14)
                plt.xlabel('Time (min)')
                plt.ylabel('PH')
            
                plt.subplot(313)#create plot of time vs. DO
                plt.plot(t1,DO1, color = 'g')
                plt.xlabel('Time (min)')
                plt.ylabel('DO (ppm)')
        
                figure1Active = True
            else:
                plt.subplot(311)  # Create plot of time vs. Temp 
                if (i1 > 100):
                    plt.plot(t1[(i1-50):i1],T1[(i1-50):i1],color = 'r')
                else:
                    plt.plot(t1,T1,color ='r')
                            
                plt.subplot(312)  # Create plot of time vs. Temp 
                if (i1 > 100):
                    plt.plot(t1[(i1-50):i1],PH1[(i1-50):i1],color = 'b')
                else:
                    plt.plot(t1,PH1,color ='b')
                plt.subplot(313)  # Create plot of time vs. Temp 
                if (i1 > 100):
                    plt.plot(t1[(i1-50):i1],DO1[(i1-50):i1],color = 'g')
                else:
                    plt.plot(t1,DO1,color ='g')
                plt.draw()  # Update plot
                plt.pause(0.05)

            count1 = count1+1  # Increase index
    if (i1 > 1):
        if (fTemp1 < T1[(i1-1)]):
            ser1.write(('PT\n').encode())
        else:
            ser1.write(('PF\n').encode())

    if (ferm2 == True):
        
        ser2.write(('F\n').encode())  # Let arduino know to send F1 data
        time.sleep(0.5)
        data = ((ser2.readline()).decode()).strip('\n')  # Readline from buffer and remove newline
        print (data)
        if (str(data[0:2]) == "F2"):
            time.sleep(0.1)
            count = 0  # Create counter for data collection
            while (count < 4):#while all data has not been collected
                data = ((ser2.readline()).decode()).strip('\n')#readline from buffer and remove newline
                print(data)
            
                if (str(data[0:2]) =="ti" and tflag == False):
                    t2.append(float(data[2:]))
                    count = count + 1
                    tflag = True
            
                if (str(data[0:2]) == "PH" and PHflag == False):
                    try:
                        strCount = 0
                        for x in range (0,len(data)):
                            check = data[x]
                            if (check == "."):
                                strCount = strCount + 1
                                strPlace = x
                        if (strCount > 1):
                            PH2.append(float((data[(x-1):]).strip('\r')))
                        else:
                            PH2.append(float((data[2:]).strip('\r')))
                    except:
                        PH2.append(len(PH2)-1)
                        #PH1.append(float((data[2:5]).strip('\r')))
                    count = count + 1
                    PHflag = True
                
                if (str(data[0:2]) == "DO" and DOflag == False):
                    try:
                        DO2.append(float(data[2:]))
                    except:
                        DO2.append(0)
                    count = count + 1
                    DOflag = True
                
                if (str(data[0:2]) == "Te" and Tflag == False):
                    T2.append(float(data[2:]))
                    count = count + 1
                    Tflag = True
                
            #Reset Flags
            PHflag = False
            DOflag = False
            tflag = False
            Tflag = False
            
            #Record Data into Excel file
            d = ws1.cell(row = 2+i2, column = 5)#call cell
            d.value = t2[i2] #set cell value
            d = ws1.cell(row = 2+i2, column = 6)#call cell
            d.value = T2[i2] #set cell value
            d = ws1.cell(row = 2+i2, column = 7)#call cell
            d.value = PH2[i2] #set cell value
            d = ws1.cell(row = 2+i2, column = 8)#call cell
            d.value = DO2[i2] #set cell value
            i2 = i2 + 1
            saveCount = saveCount + 1
            if (saveCount == 10):
                wb.save(filename = fileName)
                saveCount = 0
            
            if (figure2Active == False):#define first figure
                plt.ion()#create interactive
                plt.figure(2)
                plt.subplot(311)#create plot of time vs. Temp    
                plt.title('Fermenter 2')
                plt.plot(t2,T2, color = 'r')
                plt.ylim(0,30)
                plt.xlabel('Time (min)')
                plt.ylabel('Temp (*C)')
            
                plt.subplot(312)#create plot of time vs. PH
                plt.plot(t2,PH2, color = 'b')
                plt.ylim(0,14)
                plt.xlabel('Time (min)')
                plt.ylabel('PH')
            
                plt.subplot(313)#create plot of time vs. DO
                plt.plot(t2,DO2, color = 'g')
                plt.xlabel('Time (min)')
                plt.ylabel('DO (ppm)')
        
                figure2Active = True
            else:
                plt.subplot(311)#create plot of time vs. Temp 
                if (i2 > 100):
                    plt.plot(t2[(i2-50):i2],T2[(i2-50):i2],color = 'r')
                else:
                    plt.plot(t2,T2,color ='r')
                            
                plt.subplot(312)#create plot of time vs. Temp 
                if (i2 > 100):
                    plt.plot(t2[(i2-50):i2],PH2[(i2-50):i2],color = 'b')
                else:
                    plt.plot(t2,PH2,color ='b')
                plt.subplot(313)#create plot of time vs. Temp 
                if (i2 > 100):
                    plt.plot(t2[(i2-50):i2],DO2[(i2-50):i2],color = 'g')
                else:
                    plt.plot(t2,DO2,color ='g')
                plt.draw()#update plot
                plt.pause(0.05)
                    #plt.show()
            count2 = count2+1#increase index
    if (i2 > 1):
        if (fTemp2 < T2[(i2-1)]):
            ser2.write(('PT\n').encode())
        else:
            ser2.write(('PF\n').encode())

    if (ferm3 == True):
        ser3.write(('F\n').encode())#let arduino know to send F1 data
        time.sleep(0.5)
        data = ((ser3.readline()).decode()).strip('\n')#readline from buffer and remove newline
        print (data)
        if (str(data[0:2]) == "F3"):
            time.sleep(0.1)
            count = 0#create counter for data collection
            while (count < 4):#while all data has not been collected
                data = ((ser3.readline()).decode()).strip('\n')#readline from buffer and remove newline
                print(data)
            
                if (str(data[0:2]) =="ti" and tflag == False):
                    t3.append(float(data[2:]))
                    count = count + 1
                    tflag = True
            
                if (str(data[0:2]) == "PH" and PHflag == False):
                    try:
                        strCount = 0
                        for x in range (0,len(data)):
                            check = data[x]
                            if (check == "."):
                                strCount = strCount + 1
                                strPlace = x
                        if (strCount > 1):
                            PH3.append(float((data[(x-1):]).strip('\r')))
                        else:
                            PH3.append(float((data[2:]).strip('\r')))
                    except:
                        PH3.append(len(PH3)-1)
                        #PH1.append(float((data[2:5]).strip('\r')))
                    count = count + 1
                    PHflag = True
                
                if (str(data[0:2]) == "DO" and DOflag == False):
                    try:
                        DO3.append(float(data[2:]))
                    except:
                        DO3.append(0)
                    count = count + 1
                    DOflag = True
                
                if (str(data[0:2]) == "Te" and Tflag == False):
                    T3.append(float(data[2:]))
                    count = count + 1
                    Tflag = True
                
            #Reset Flags
            PHflag = False
            DOflag = False
            tflag = False
            Tflag = False
            
            #Record Data into Excel file
            d = ws1.cell(row = 2+i2, column = 9)#call cell
            d.value = t3[i3] #set cell value
            d = ws1.cell(row = 2+i2, column = 10)#call cell
            d.value = T3[i3] #set cell value
            d = ws1.cell(row = 2+i2, column = 11)#call cell
            d.value = PH3[i3] #set cell value
            d = ws1.cell(row = 2+i2, column = 12)#call cell
            d.value = DO3[i3] #set cell value
            i3 = i3 + 1
            saveCount = saveCount + 1
            if (saveCount == 10):
                wb.save(filename = fileName)
                saveCount = 0
            
            if (figure3Active == False):#define first figure
                plt.ion()#create interactive
                plt.figure(3)
                plt.subplot(311)#create plot of time vs. Temp    
                plt.title('Fermenter 3')
                plt.plot(t3,T3, color = 'r')
                plt.ylim(0,30)
                plt.xlabel('Time (min)')
                plt.ylabel('Temp (*C)')
            
                plt.subplot(312)#create plot of time vs. PH
                plt.plot(t3,PH3, color = 'b')
                plt.ylim(0,14)
                plt.xlabel('Time (min)')
                plt.ylabel('PH')
            
                plt.subplot(313)#create plot of time vs. DO
                plt.plot(t3,DO3, color = 'g')
                plt.xlabel('Time (min)')
                plt.ylabel('DO (ppm)')
        
                figure3Active = True
            else:
                plt.subplot(311)#create plot of time vs. Temp 
                if (i3 > 100):
                    plt.plot(t3[(i3-50):i3],T3[(i3-50):i3],color = 'r')
                else:
                    plt.plot(t3,T3,color ='r')
                            
                plt.subplot(312)#create plot of time vs. Temp 
                if (i3 > 100):
                    plt.plot(t3[(i3-50):i3],PH3[(i2-50):i3],color = 'b')
                else:
                    plt.plot(t3,PH3,color ='b')
                plt.subplot(313)#create plot of time vs. Temp 
                if (i3 > 100):
                    plt.plot(t3[(i3-50):i3],DO3[(i3-50):i3],color = 'g')
                else:
                    plt.plot(t3,DO3,color ='g')
                plt.draw()#update plot
                plt.pause(0.05)
                    #plt.show()
            count3 = count3+1#increase index
    if (i3 > 1):
        if (fTemp3 < T3[(i3-1)]):
            ser3.write(('PT\n').encode())
        else:
            ser3.write(('PF\n').encode())

    
    start_time = time.time()
    elapsed_time = 0
    while(elapsed_time < 5):
        elapsed_time = time.time() - start_time
        #Wait 1 minute before next iteration
        try:
            if (keyboard.is_pressed('q')):#ifa key is pressed
                keyPressed = True
                while(keyPressed == True):
                    reply = input("please input a command\n")
                    print(reply)
                    if (str(reply)=="ct1"):
                        ser1.write(('CT\n').encode())
                        fTemp1 = fermTemp(1)
                        keyPressed = False
                    
                    if (str(reply)=="ct2"):
                        ser2.write(('CT\n').encode())
                        fTemp2 = fermTemp(2)
                        keyPressed = False
                    
                    if (str(reply)=="ct3"):
                        ser3.write(('CT\n').encode())
                        fTemp3 = fermTemp(3)
                        keyPressed = False
                    
                    if (str(reply)=="End"):
                        if (ferm1==True):
                            ser1.write(('E\n').encode())
                            time.sleep(0.5)
                            ser1.close()
                            
                        if (ferm2==True):
                            ser2.write(('E\n').encode())
                            time.sleep(0.5)
                            ser2.close()
                            
                        if (ferm3==True):
                            ser3.write(('E\n').encode())
                            time.sleep(0.5)
                            ser3.close()
                            
                        fermAct = False
                        keyPressed = False
                    else:
                        print("error")
        except:
            pass            
    print ("next iteration")

